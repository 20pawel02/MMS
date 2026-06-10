"""
Projekt zaliczeniowy – Wersja 2: Korelacja i regresja
Zbiór: Heart Disease Database (Cleveland, UCI)
Zmienna objaśniana Y: oldpeak (depresja ST w teście wysiłkowym)
"""

from __future__ import annotations

import os
import urllib.request
from itertools import combinations
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from fpdf import FPDF
from openpyxl import Workbook
from scipy.stats import pearsonr, t as t_dist
from sklearn.linear_model import LinearRegression

# =============================================================================
# KONFIGURACJA – uzupełnij przed oddaniem
# =============================================================================
NAZWISKA_ZESPOLU = "Nazwisko1_Nazwisko2"
PODZIAL_OBOWIAZKOW = (
    "Podział obowiązków:\n"
    "Osoba 1: wczytanie danych, model pełny, metoda Pawłowskiego, raport PDF.\n"
    "Osoba 2: eliminacja zmiennych, modele jednej zmiennej, arkusz Excel, wykresy."
)

KATALOG_PROJEKTU = Path(__file__).resolve().parent
PLIK_DANYCH = KATALOG_PROJEKTU / "data" / "processed.cleveland.data"
KATALOG_WYNIKOW = KATALOG_PROJEKTU / "wyniki"

NAZWY_KOLUMN = [
    "age", "sex", "cp", "trestbps", "chol", "fbs", "restecg",
    "thalach", "exang", "oldpeak", "slope", "ca", "thal", "num",
]
Y_NAZWA = "oldpeak"
X_KANDYDACI = ["age", "trestbps", "chol", "thalach", "ca", "cp"]

R_KRYT = 0.5
PAWLOWSKI_K = 2
POZIOMY_ALFA = [0.01, 0.05, 0.1]
V_QUASI = 0.10  # usuwanie quasi-stałych: CV > 10%

# Wartości do prognozy (profil „typowego” pacjenta z próby – mediany)
PROGNOZA_WIELOMA = {}
PROGNOZA_JEDNA_X = {}  # uzupełniane po wczytaniu danych


# =============================================================================
# WCZYTYWANIE DANYCH
# =============================================================================
def wczytaj_cleveland(sciezka: Path) -> tuple[np.ndarray, list[str]]:
    wiersze = []
    with open(sciezka, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            czesci = [p.strip() for p in line.split(",")]
            wiersz = []
            for p in czesci:
                if p in ("?", "", "NA", "nan"):
                    wiersz.append(np.nan)
                else:
                    wiersz.append(float(p))
            wiersze.append(wiersz)
    return np.asarray(wiersze), NAZWY_KOLUMN


def wyodrebnij_zmienne(
    macierz: np.ndarray, nazwy: list[str], kolumny: list[str]
) -> dict[str, np.ndarray]:
    indeks = {n: i for i, n in enumerate(nazwy)}
    return {k: macierz[:, indeks[k]] for k in kolumny}


def usun_braki(slownik: dict[str, np.ndarray]) -> dict[str, np.ndarray]:
    macierz = np.column_stack([slownik[k] for k in slownik])
    maska = np.all(np.isfinite(macierz), axis=1)
    return {k: v[maska] for k, v in slownik.items()}


def wspolczynnik_zmiennosci(x: np.ndarray) -> float:
    srednia = np.mean(x)
    if abs(srednia) < 1e-12:
        return 0.0
    return abs(np.std(x, ddof=1) / srednia) * 100


def usun_quasistale(zmienne: list[np.ndarray], prog: float) -> list[np.ndarray]:
    return [z for z in zmienne if wspolczynnik_zmiennosci(z) > prog * 100]


def utworz_wektor_korelacji(zmienne: list[np.ndarray], y: np.ndarray) -> list[float]:
    return [abs(pearsonr(x, y)[0]) for x in zmienne]


def utworz_macierz_korelacji(zmienne: list[np.ndarray], bezwzgledna: bool = True) -> list[list[float]]:
    macierz = []
    for x in zmienne:
        wiersz = []
        for z in zmienne:
            r = pearsonr(x, z)[0]
            wiersz.append(abs(r) if bezwzgledna else r)
        macierz.append(wiersz)
    return macierz


def eliminacja_nadmiarowych_zmiennych(
    wektor_korelacji: list[float], macierz_korelacji: list[list[float]], r_kryt: float
) -> list[int]:
    wektor = np.asarray(wektor_korelacji)
    r_mat = np.asarray(macierz_korelacji)
    kandydaci = [i for i, r in enumerate(wektor) if abs(r) >= r_kryt]
    wybrane = []
    while kandydaci:
        najlepsza = max(kandydaci, key=lambda i: abs(wektor[i]))
        wybrane.append(najlepsza)
        print(f"  wybrane: {[i + 1 for i in wybrane]}, kandydaci przed redukcja: {[i + 1 for i in kandydaci]}")
        kandydaci = [
            i for i in kandydaci if i != najlepsza and r_mat[najlepsza, i] <= r_kryt
        ]
        print(f"  kandydaci po redukcji: {[i + 1 for i in kandydaci]}")
    if not wybrane and len(wektor) > 0:
        wybrane = [int(np.argmax(np.abs(wektor)))]
    return [i + 1 for i in wybrane]


def metoda_pawlowskiego(macierz_korelacji: list[list[float]], k: int) -> tuple[list[int], float]:
    r_mat = np.asarray(macierz_korelacji)
    n = r_mat.shape[0]
    max_det = -np.inf
    najlepszy = None
    for idx in combinations(range(n), k):
        pod = r_mat[np.ix_(idx, idx)]
        det = np.linalg.det(pod)
        if det > max_det:
            max_det = det
            najlepszy = idx
    return [i + 1 for i in najlepszy], float(max_det)


def regresja_wieloraka(x_list: list[np.ndarray], y: np.ndarray) -> LinearRegression:
    x_mat = np.column_stack(x_list)
    model = LinearRegression().fit(x_mat, y)
    return model


def prognoza_wieloraka(model: LinearRegression, x_wartosci: list[float]) -> float:
    return float(model.predict([x_wartosci])[0])


def test_istotnosci_korelacji(r: float, n: int, alfa_lista: list[float]) -> dict:
    """H0: rho = 0; test t dla współczynnika korelacji Pearsona."""
    if n < 3 or abs(r) >= 1:
        return {"r": r, "t": np.nan, "df": n - 2, "wyniki": []}
    t_stat = r * np.sqrt((n - 2) / (1 - r**2))
    df = n - 2
    wyniki = []
    for alfa in alfa_lista:
        t_kryt = float(t_dist.ppf(1 - alfa / 2, df))
        odrzuc = abs(t_stat) > t_kryt
        wyniki.append(
            {
                "alfa": alfa,
                "t_stat": t_stat,
                "t_kryt": t_kryt,
                "odrzuc_h0": odrzuc,
                "tekst": "Odrzucamy H0" if odrzuc else "Brak podstaw do odrzucenia H0",
            }
        )
    return {"r": r, "t": t_stat, "df": df, "wyniki": wyniki}


def modele_jednej_zmiennej(x: np.ndarray, y: np.ndarray, nazwa_x: str, katalog: Path):
    """Pkt 9: liniowy, nieliniowe, wielomianowy, wybór najlepszego."""
    n = len(y)
    wyniki = {}
    x_prog = float(np.median(x))
    y_prog = None
    najlepszy_klucz = None
    najlepsze_r2 = -np.inf

    # a) liniowy
    model_lin = LinearRegression().fit(x.reshape(-1, 1), y)
    r2_lin = model_lin.score(x.reshape(-1, 1), y)
    r_pearson, p_pearson = pearsonr(x, y)
    test = test_istotnosci_korelacji(r_pearson, n, POZIOMY_ALFA)
    wyniki["liniowy"] = {
        "r2": r2_lin,
        "model": model_lin,
        "r": r_pearson,
        "p": p_pearson,
        "test": test,
        "prognoza_x": x_prog,
        "prognoza_y": float(model_lin.predict([[x_prog]])[0]),
        "opis": f"Y = {model_lin.intercept_:.4f} + {model_lin.coef_[0]:.4f}*{nazwa_x}",
    }

    # b) nieliniowe
    propozycje = {}
    if np.all(x > 0):
        x_log = np.log(x).reshape(-1, 1)
        m = LinearRegression().fit(x_log, y)
        propozycje["logarytmiczny"] = (m, m.score(x_log, y), f"Y = a + b*ln({nazwa_x})")

    if np.all(y > 0) and np.all(x > 0):
        y_log = np.log(y)
        m = LinearRegression().fit(x.reshape(-1, 1), y_log)
        propozycje["wykładniczy"] = (m, m.score(x.reshape(-1, 1), y_log), f"ln(Y) = a + b*{nazwa_x}")

    if np.all(x != 0):
        x_inv = (1 / x).reshape(-1, 1)
        m = LinearRegression().fit(x_inv, y)
        propozycje["odwrotnościowy"] = (m, m.score(x_inv, y), f"Y = a + b*(1/{nazwa_x})")

    wyniki["nieliniowe"] = {}
    for klucz, (m, r2, opis) in propozycje.items():
        if klucz == "logarytmiczny":
            y_hat = m.predict(np.array([[np.log(x_prog)]]))
        elif klucz == "wykładniczy":
            y_hat = np.exp(m.predict([[x_prog]]))
        elif klucz == "odwrotnościowy":
            y_hat = m.predict([[1 / x_prog]])
        else:
            y_hat = [np.nan]
        wyniki["nieliniowe"][klucz] = {"r2": r2, "opis": opis, "prognoza_y": float(y_hat[0])}

    # c) wielomian stopnia 2
    x_wiel = np.column_stack([x, x**2])
    model_wiel = LinearRegression().fit(x_wiel, y)
    r2_wiel = model_wiel.score(x_wiel, y)
    a0, a1, a2 = model_wiel.intercept_, model_wiel.coef_[0], model_wiel.coef_[1]
    y_hat_wiel = a0 + a1 * x_prog + a2 * x_prog**2
    wyniki["wielomian"] = {
        "r2": r2_wiel,
        "model": model_wiel,
        "prognoza_y": float(y_hat_wiel),
        "opis": f"Y = {a0:.4f} + {a1:.4f}*{nazwa_x} + {a2:.4f}*{nazwa_x}^2",
    }

    # d) najlepszy model (max R²)
    kandydaci = {"liniowy": r2_lin, "wielomian": r2_wiel}
    for klucz, d in wyniki["nieliniowe"].items():
        kandydaci[f"nieliniowy_{klucz}"] = d["r2"]
    najlepszy_klucz = max(kandydaci, key=kandydaci.get)
    najlepsze_r2 = kandydaci[najlepszy_klucz]

    if najlepszy_klucz == "liniowy":
        y_prog = wyniki["liniowy"]["prognoza_y"]
    elif najlepszy_klucz == "wielomian":
        y_prog = wyniki["wielomian"]["prognoza_y"]
    else:
        _, typ = najlepszy_klucz.split("_", 1)
        y_prog = wyniki["nieliniowe"][typ]["prognoza_y"]

    wyniki["najlepszy"] = {
        "klucz": najlepszy_klucz,
        "r2": najlepsze_r2,
        "prognoza_x": x_prog,
        "prognoza_y": y_prog,
    }

    # wykres
    fig, ax = plt.subplots(figsize=(8, 4))
    x_sort = np.sort(x)
    ax.scatter(x, y, color="crimson", alpha=0.6, label="dane")
    if najlepszy_klucz == "liniowy":
        y_line = model_lin.predict(x_sort.reshape(-1, 1))
    elif najlepszy_klucz == "wielomian":
        y_line = model_wiel.predict(np.column_stack([x_sort, x_sort**2]))
    elif najlepszy_klucz.startswith("nieliniowy_"):
        typ = najlepszy_klucz.replace("nieliniowy_", "")
        if typ == "logarytmiczny":
            y_line = propozycje["logarytmiczny"][0].predict(np.log(x_sort).reshape(-1, 1))
        elif typ == "wykładniczy":
            y_line = np.exp(propozycje["wykładniczy"][0].predict(x_sort.reshape(-1, 1)))
        else:
            y_line = propozycje["odwrotnościowy"][0].predict((1 / x_sort).reshape(-1, 1))
    else:
        y_line = model_lin.predict(x_sort.reshape(-1, 1))
    ax.plot(x_sort, y_line, color="navy", label=f"najlepszy: {najlepszy_klucz}")
    ax.axvline(x_prog, color="gray", linestyle="--", alpha=0.5)
    ax.scatter([x_prog], [y_prog], color="green", s=80, zorder=5, label="prognoza")
    ax.set_xlabel(nazwa_x)
    ax.set_ylabel(Y_NAZWA)
    ax.set_title(f"Analiza jednej zmiennej: {nazwa_x} → {Y_NAZWA}")
    ax.legend()
    fig.tight_layout()
    sciezka_wykresu = katalog / f"wykres_{nazwa_x}.png"
    fig.savefig(sciezka_wykresu, dpi=120)
    plt.close(fig)
    wyniki["wykres"] = str(sciezka_wykresu)
    return wyniki


# =============================================================================
# PDF (polskie znaki – DejaVu)
# =============================================================================
def pobierz_czcionki(katalog: Path) -> tuple[str, str]:
    regular = katalog / "DejaVuSans.ttf"
    bold = katalog / "DejaVuSans-Bold.ttf"
    url_r = "https://raw.githubusercontent.com/matplotlib/matplotlib/main/lib/matplotlib/mpl-data/fonts/ttf/DejaVuSans.ttf"
    url_b = "https://raw.githubusercontent.com/matplotlib/matplotlib/main/lib/matplotlib/mpl-data/fonts/ttf/DejaVuSans-Bold.ttf"
    if not regular.exists():
        urllib.request.urlretrieve(url_r, regular)
    if not bold.exists():
        urllib.request.urlretrieve(url_b, bold)
    return str(regular), str(bold)


def generuj_pdf(
    sciezka: Path,
    opis: dict,
    modele: dict,
    analizy_jedna: dict,
    font_regular: str,
    font_bold: str,
):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_font("DejaVu", "", font_regular)
    pdf.add_font("DejaVu", "B", font_bold)

    szer = 190

    def naglowek(tekst: str):
        pdf.set_font("DejaVu", "B", 12)
        pdf.set_x(10)
        pdf.multi_cell(szer, 8, tekst)
        pdf.set_font("DejaVu", "", 10)
        pdf.ln(2)

    def tekst(t: str):
        pdf.set_font("DejaVu", "", 10)
        pdf.set_x(10)
        pdf.multi_cell(szer, 6, t)

    pdf.add_page()
    naglowek(f"Projekt MMS – Korelacja i regresja – {NAZWISKA_ZESPOLU}")
    tekst(
        "Temat: modelowanie depresji odcinka ST (oldpeak) u pacjentów z bazy Cleveland "
        "(Heart Disease Database, UCI Machine Learning Repository)."
    )
    tekst(opis["cel_badawczy"])
    pdf.ln(3)

    naglowek("1. Opis zestawu danych")
    tekst(opis["opis_danych"])
    tekst(opis["tabela_cv"])
    pdf.ln(3)

    naglowek("2. Model regresji wielorakiej – wszystkie zmienne objaśniające")
    tekst(modele["pelny"]["tekst"])
    pdf.ln(3)

    naglowek("3. Metoda Pawłowskiego (k = 2)")
    tekst(modele["pawlowski"]["tekst"])
    pdf.ln(3)

    naglowek("4. Redukcja zmiennych – analiza współczynników korelacji")
    tekst(modele["eliminacja"]["tekst"])
    pdf.ln(3)

    naglowek("5. Model po redukcji (eliminacja nadmiarowych)")
    tekst(modele["po_eliminacji"]["tekst"])
    pdf.ln(3)

    for nazwa, analiza in analizy_jedna.items():
        pdf.add_page()
        naglowek(f"6. Analiza pojedynczej zmiennej: {nazwa}")
        lin = analiza["liniowy"]
        tekst(f"a) Model liniowy: {lin['opis']}, R² = {lin['r2']:.4f}")
        tekst(f"   H0: ρ = 0; H1: ρ ≠ 0. r = {lin['r']:.4f}, t = {lin['test']['t']:.4f}, df = {lin['test']['df']}")
        for w in lin["test"]["wyniki"]:
            tekst(f"   α = {w['alfa']}: t_kryt = {w['t_kryt']:.4f} → {w['tekst']}")
        tekst("b) Modele nieliniowe:")
        for klucz, d in analiza["nieliniowe"].items():
            tekst(f"   {klucz}: {d['opis']}, R² = {d['r2']:.4f}")
        wiel = analiza["wielomian"]
        tekst(f"c) Wielomian: {wiel['opis']}, R² = {wiel['r2']:.4f}")
        naj = analiza["najlepszy"]
        tekst(
            f"d) Najlepszy model: {naj['klucz']}, R² = {naj['r2']:.4f}. "
            f"Prognoza dla {nazwa} = {naj['prognoza_x']:.2f}: Y = {naj['prognoza_y']:.4f}"
        )
        if os.path.exists(analiza["wykres"]):
            pdf.image(analiza["wykres"], w=170)

    pdf.add_page()
    naglowek("7. Podsumowanie")
    tekst(modele["podsumowanie"])
    pdf.ln(5)
    naglowek("Podział obowiązków")
    tekst(PODZIAL_OBOWIAZKOW)

    pdf.output(str(sciezka))


# =============================================================================
# EXCEL
# =============================================================================
def zapisz_excel(
    sciezka: Path,
    dane: dict,
    modele: dict,
    analizy: dict,
    nazwy_x: list[str],
    macierz_korelacji: list[list[float]],
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dane"
    ws.append(NAZWY_KOLUMN)
    for row in dane["macierz_pelna"]:
        ws.append(list(row))

    ws2 = wb.create_sheet("CV_i_wybor")
    ws2.append(["Zmienna", "Średnia", "Odch.std", "CV%", "Wybrana"])
    for wiersz in dane["statystyki_cv"]:
        ws2.append(wiersz)

    def arkusz_modelu(nazwa: str, naglowek: str, wiersze: list):
        sh = wb.create_sheet(nazwa)
        sh.append([naglowek])
        for w in wiersze:
            sh.append(w if isinstance(w, list) else [w])

    sh_r = wb.create_sheet("Macierz_korelacji")
    sh_r.append([""] + nazwy_x)
    for i, nazwa in enumerate(nazwy_x):
        wiersz = [round(macierz_korelacji[i][j], 4) for j in range(len(nazwy_x))]
        sh_r.append([nazwa] + wiersz)

    arkusz_modelu("Model_pelny", "Regresja wieloraka – pełny", modele["pelny"]["excel"])
    arkusz_modelu("Pawlowski", "Metoda Pawłowskiego", modele["pawlowski"]["excel"])
    arkusz_modelu("Eliminacja", "Eliminacja zmiennych", modele["eliminacja"]["excel"])
    arkusz_modelu("Model_redukcja", "Po eliminacji", modele["po_eliminacji"]["excel"])

    for nazwa, analiza in analizy.items():
        sh = wb.create_sheet(f"Jedna_{nazwa}"[:31])
        sh.append([f"Zmienna objaśniająca: {nazwa}"])
        sh.append(["Model", "R2", "Uwagi"])
        sh.append(["liniowy", analiza["liniowy"]["r2"], analiza["liniowy"]["opis"]])
        for klucz, d in analiza["nieliniowe"].items():
            sh.append([klucz, d["r2"], d["opis"]])
        sh.append(["wielomian", analiza["wielomian"]["r2"], analiza["wielomian"]["opis"]])
        sh.append(["NAJLEPSZY", analiza["najlepszy"]["r2"], analiza["najlepszy"]["klucz"]])

    wb.save(sciezka)


# =============================================================================
# PROGRAM GŁÓWNY
# =============================================================================
def main():
    KATALOG_WYNIKOW.mkdir(parents=True, exist_ok=True)
    print("=" * 60)
    print("Projekt MMS – Korelacja i regresja (Wersja 2)")
    print("=" * 60)

    macierz, nazwy = wczytaj_cleveland(PLIK_DANYCH)
    print(f"Wczytano {macierz.shape[0]} rekordów, {macierz.shape[1]} kolumn.")

  # Wybór zmiennych z CV > 10%
    wszystkie = wyodrebnij_zmienne(macierz, nazwy, NAZWY_KOLUMN)
    stat_cv = []
    wybrane_x_nazwy = []
    for nazwa in X_KANDYDACI:
        if nazwa == Y_NAZWA:
            continue
        x = wszystkie[nazwa]
        x_ok = x[np.isfinite(x)]
        cv = wspolczynnik_zmiennosci(x_ok)
        wybrana = cv > V_QUASI * 100
        stat_cv.append([nazwa, round(float(np.mean(x_ok)), 2), round(float(np.std(x_ok, ddof=1)), 2), round(cv, 2), wybrana])
        if wybrana:
            wybrane_x_nazwy.append(nazwa)

    if len(wybrane_x_nazwy) < 4:
        raise ValueError(
            f"Wymagane co najmniej 4 zmienne z CV>10%. Wybrano: {wybrane_x_nazwy}. "
            "Rozszerz listę X_KANDYDACI w konfiguracji."
        )

    zmienne_do_analizy = [Y_NAZWA] + wybrane_x_nazwy
    dane = wyodrebnij_zmienne(macierz, nazwy, zmienne_do_analizy)
    dane = usun_braki(dane)
    n = len(dane[Y_NAZWA])
    y = dane[Y_NAZWA]
    print(f"Po usunięciu braków: n = {n}")
    print(f"Zmienne objaśniające (CV > 10%): {wybrane_x_nazwy}")

    for nazwa in wybrane_x_nazwy:
        PROGNOZA_WIELOMA[nazwa] = float(np.median(dane[nazwa]))
        PROGNOZA_JEDNA_X[nazwa] = PROGNOZA_WIELOMA[nazwa]

    x_lista = [dane[n] for n in wybrane_x_nazwy]
    x_nazwy = wybrane_x_nazwy

    cel_badawczy = (
        "Cel badawczy: zbadanie wpływu parametrów klinicznych pacjenta "
        f"({', '.join(x_nazwy)}) na nasilenie depresji odcinka ST w teście wysiłkowym "
        f"({Y_NAZWA}), z wykorzystaniem regresji wielorakiej oraz metod redukcji zmiennych."
    )
    print(cel_badawczy)

    opis_danych = (
        f"Baza Cleveland (Heart Disease Database): {n} obserwacji po usunięciu braków. "
        f"Zmienna objaśniana: {Y_NAZWA}. Zmienne objaśniające: {', '.join(x_nazwy)}. "
        "Atrybuty numeryczne; braki oznaczone w oryginale znakiem '?'."
    )
    tabela_cv = "Współczynniki zmienności:\n" + "\n".join(
        f"  {r[0]}: średnia={r[1]}, s={r[2]}, CV={r[3]}% {'[WYBRANA]' if r[4] else ''}"
        for r in stat_cv
    )

    # --- Model pełny ---
    model_pelny = regresja_wieloraka(x_lista, y)
    r2_pelny = model_pelny.score(np.column_stack(x_lista), y)
    x_prog_pelny = [PROGNOZA_WIELOMA[n] for n in x_nazwy]
    y_prog_pelny = prognoza_wieloraka(model_pelny, x_prog_pelny)
    wsp_pelny = " ".join(f"{c:+.4f}*{n}" for c, n in zip(model_pelny.coef_, x_nazwy))
    tekst_pelny = (
        f"Y = {model_pelny.intercept_:.4f} {wsp_pelny}\n"
        f"R^2 = {r2_pelny:.4f} – model wyjaśnia {r2_pelny*100:.1f}% wariancji {Y_NAZWA}.\n"
        f"Prognoza dla profilu (mediany): {dict(zip(x_nazwy, x_prog_pelny))} -> Y = {y_prog_pelny:.4f}"
    )
    print("\n--- Model pełny ---")
    print(tekst_pelny)

    # --- Pawłowski ---
    macierz_p = utworz_macierz_korelacji(x_lista, bezwzgledna=False)
    indeksy_paw, det_paw = metoda_pawlowskiego(macierz_p, PAWLOWSKI_K)
    nazwy_paw = [x_nazwy[i - 1] for i in indeksy_paw]
    x_paw = [dane[n] for n in nazwy_paw]
    model_paw = regresja_wieloraka(x_paw, y)
    r2_paw = model_paw.score(np.column_stack(x_paw), y)
    x_prog_paw = [PROGNOZA_WIELOMA[n] for n in nazwy_paw]
    y_prog_paw = prognoza_wieloraka(model_paw, x_prog_paw)
    tekst_pawlowski = (
        f"Wybrane zmienne (numery 1..p): {indeksy_paw} -> {nazwy_paw}, det(R) = {det_paw:.6f}\n"
        f"R^2 = {r2_paw:.4f}. Prognoza: {y_prog_paw:.4f}"
    )
    print("\n--- Pawłowski ---")
    print(tekst_pawlowski)

    # --- Eliminacja nadmiarowych ---
    zmienne_z = usun_quasistale(x_lista, V_QUASI)
    nazwy_z = [x_nazwy[i] for i, z in enumerate(x_lista) if wspolczynnik_zmiennosci(z) > V_QUASI * 100]
    wektor = utworz_wektor_korelacji(zmienne_z, y)
    macierz_abs = utworz_macierz_korelacji(zmienne_z, bezwzgledna=True)
    print("Eliminacja nadmiarowych (kroki):")
    wybrane_idx = eliminacja_nadmiarowych_zmiennych(wektor, macierz_abs, R_KRYT)
    nazwy_elim = [nazwy_z[i - 1] for i in wybrane_idx]
    x_elim = [dane[n] for n in nazwy_elim]
    model_elim = regresja_wieloraka(x_elim, y)
    r2_elim = model_elim.score(np.column_stack(x_elim), y)
    x_prog_elim = [PROGNOZA_WIELOMA[n] for n in nazwy_elim]
    y_prog_elim = prognoza_wieloraka(model_elim, x_prog_elim)
    tekst_elim = (
        f"Po eliminacji (r_kryt = {R_KRYT}): zmienne {nazwy_elim}\n"
        f"R^2 = {r2_elim:.4f}. Prognoza: {y_prog_elim:.4f}"
    )
    print("\n--- Eliminacja korelacji ---")
    print(tekst_elim)

    # --- Pkt 9: dla każdej zmiennej z Pawłowskiego ---
    analizy_jedna = {}
    for nazwa in nazwy_paw:
        print(f"\n--- Analiza jednej zmiennej: {nazwa} ---")
        analizy_jedna[nazwa] = modele_jednej_zmiennej(
            dane[nazwa], y, nazwa, KATALOG_WYNIKOW
        )
        naj = analizy_jedna[nazwa]["najlepszy"]
        print(f"  Najlepszy: {naj['klucz']}, R^2={naj['r2']:.4f}, prognoza Y={naj['prognoza_y']:.4f}")

    podsumowanie = (
        f"Porównanie R^2: model pelny {r2_pelny:.4f}, Pawlowski {r2_paw:.4f}, "
        f"po eliminacji {r2_elim:.4f}. Redukcja zmiennych obniża złożoność; "
        f"pełny model zwykle ma najwyższe dopasowanie na próbie uczącej. "
        f"Wnioski kliniczne należy formułować ostrożnie – korelacja nie implikuje przyczynowości."
    )

    modele = {
        "pelny": {
            "tekst": tekst_pelny,
            "excel": [
                ["R2", r2_pelny],
                ["intercept", model_pelny.intercept_],
                ["prognoza_Y", y_prog_pelny],
            ]
            + [[f"coef_{n}", c] for n, c in zip(x_nazwy, model_pelny.coef_)],
        },
        "pawlowski": {
            "tekst": tekst_pawlowski,
            "excel": [["zmienne", ", ".join(nazwy_paw)], ["det", det_paw], ["R2", r2_paw]],
        },
        "eliminacja": {
            "tekst": f"Wektor |r| z Y: {[round(v, 3) for v in wektor]}\nWybrane: {nazwy_elim}",
            "excel": [["wybrane", ", ".join(nazwy_elim)]],
        },
        "po_eliminacji": {"tekst": tekst_elim, "excel": [["R2", r2_elim], ["prognoza", y_prog_elim]]},
        "podsumowanie": podsumowanie,
    }

    opis = {
        "cel_badawczy": cel_badawczy,
        "opis_danych": opis_danych,
        "tabela_cv": tabela_cv,
    }

    dane_export = {
        "macierz_pelna": macierz,
        "statystyki_cv": stat_cv,
    }

    font_r, font_b = pobierz_czcionki(KATALOG_WYNIKOW)
    pdf_path = KATALOG_WYNIKOW / f"Raport_{NAZWISKA_ZESPOLU}.pdf"
    xlsx_path = KATALOG_WYNIKOW / f"Obliczenia_{NAZWISKA_ZESPOLU}.xlsx"

    generuj_pdf(pdf_path, opis, modele, analizy_jedna, font_r, font_b)
    zapisz_excel(xlsx_path, dane_export, modele, analizy_jedna, x_nazwy, macierz_p)


if __name__ == "__main__":
    main()
